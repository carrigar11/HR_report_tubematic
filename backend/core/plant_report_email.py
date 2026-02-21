"""
Daily Plant Report (Previous day) email: send Excel + image to recipients at configured time.
Recipients and send time are stored in DB (PlantReportRecipient, SystemSetting).
Attachment: Plant Report sheet (Excel) + PNG image for easy viewing on mobile.
"""
import logging
from datetime import date
from io import BytesIO

from django.core.mail import EmailMultiAlternatives, get_connection
from email.mime.image import MIMEImage
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from .email_smtp import get_active_smtp_config
from .google_sheets_sync import _build_sheet3_data
from .models import PlantReportRecipient, SystemSetting

logger = logging.getLogger(__name__)

# Column indices for color scale (0-based): 7=Avg Salary/hr, 8=Absenteeism %, 10=OT bonus hrs, 11=OT bonus rs
_COLOR_SCALE_COLS = (7, 8, 10, 11)


def _value_to_float(val):
    """Convert cell value to float for color scale; return None if not numeric."""
    if val is None or val == '' or (isinstance(val, str) and val.strip().startswith('=')):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _color_scale_rgb(val, lo, hi):
    """Return (r,g,b) soft green->yellow->red (muted, less vibrant)."""
    if lo is None or hi is None or lo == hi:
        return (255, 252, 220)  # very light cream
    t = (val - lo) / (hi - lo) if hi > lo else 0.5
    t = max(0, min(1, t))
    if t <= 0.5:
        # soft green -> soft yellow
        u = t * 2
        return (220, 245, 200 + int(35 * (1 - u)))  # pale green to pale yellow
    else:
        # soft yellow -> soft red/coral
        u = (t - 0.5) * 2
        return (255, 245 - int(120 * u), 220 - int(150 * u))  # pale yellow to soft coral


def build_plant_report_previous_day_image():
    """
    Build Plant Report (Previous day) as a PNG image – same data as Google Sheet.
    Green->yellow->red color scale on Average Salary/hr, Absenteeism %, OT bonus (hrs), OT bonus (rs).
    Returns BytesIO (PNG).
    """
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        logger.warning('Pillow not installed; cannot generate Plant Report image')
        return None

    rows = _build_sheet3_data()
    if not rows:
        return None

    # Layout: consistent column widths so header names are visible
    cell_h = 26
    header_h = 32
    col_widths = [40, 110, 88, 80, 80, 80, 88, 88, 88, 88, 88, 88, 88]  # wider for header names
    ncols = max(len(r) for r in rows) if rows else 0
    while len(col_widths) < ncols:
        col_widths.append(88)
    col_widths = col_widths[:ncols]
    total_w = sum(col_widths) + (ncols + 1) * 1
    total_h = header_h + (len(rows) - 1) * cell_h + (len(rows) + 1) * 1

    img = Image.new('RGB', (total_w, total_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    import os
    _font_paths = [
        'arial.ttf',
        os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts', 'arial.ttf'),
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    font = font_bold = font_header = None
    for path in _font_paths:
        try:
            font = ImageFont.truetype(path, 11)
            font_bold = font
            font_header = ImageFont.truetype(path, 9)  # smaller so column names fit
            try:
                font_bold = ImageFont.truetype(
                    path.replace('arial.ttf', 'arialbd.ttf').replace('DejaVuSans.ttf', 'DejaVuSans-Bold.ttf'), 11
                )
                font_header = ImageFont.truetype(
                    path.replace('arial.ttf', 'arialbd.ttf').replace('DejaVuSans.ttf', 'DejaVuSans-Bold.ttf'), 9
                )
            except (OSError, IOError):
                pass
            break
        except (OSError, IOError):
            continue
    if font is None:
        font = font_bold = font_header = ImageFont.load_default()

    # Compute min/max per column for color scale (data rows only, numeric)
    col_mins = [None] * ncols
    col_maxs = [None] * ncols
    for ri, row in enumerate(rows):
        if ri == 0:
            continue
        for ci in _COLOR_SCALE_COLS:
            if ci < len(row):
                v = _value_to_float(row[ci])
                if v is not None:
                    if col_mins[ci] is None or v < col_mins[ci]:
                        col_mins[ci] = v
                    if col_maxs[ci] is None or v > col_maxs[ci]:
                        col_maxs[ci] = v

    def draw_cell(r, c, text, fill=(255, 255, 255), font_=font, bold=False, max_chars=14):
        x0 = 1 + sum(col_widths[:c]) + c
        y0 = 1 if r == 0 else (1 + header_h + 1 + (r - 1) * (cell_h + 1))
        w = col_widths[c] if c < len(col_widths) else 88
        h = header_h if r == 0 else cell_h
        draw.rectangle([x0, y0, x0 + w, y0 + h], fill=fill, outline=(180, 180, 180))
        tstr = str(text)
        s = tstr[:max_chars] + ('…' if len(tstr) > max_chars else '')
        # center text in cell (simple)
        bbox = draw.textbbox((0, 0), s, font=font_)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        tx = x0 + max(0, (w - tw) // 2)
        ty = y0 + max(0, (h - th) // 2)
        text_color = (255, 255, 255) if (fill[0] < 100 and fill[1] < 100) else (0, 0, 0)
        draw.text((tx, ty), s, fill=text_color, font=font_)

    # Column indices: headers = Sr No, PLANT, Total Man Hrs (3) + n_dates + 8 cols → n_dates = len(rows[0]) - 11
    n_dates = max(0, (len(rows[0]) - 11)) if rows else 0
    abs_col = 3 + n_dates + 4   # Absenteeism % column (0-based)
    present_col = 3 + n_dates
    absent_col = 3 + n_dates + 1

    for ri, row in enumerate(rows):
        for ci in range(len(row)):
            val = row[ci]
            # In total row, Absenteeism % cell is the formula – show computed number in image
            if ri == len(rows) - 1 and ri > 0 and ci == abs_col and isinstance(val, str) and val.strip().startswith('='):
                try:
                    pres = _value_to_float(row[present_col])
                    absv = _value_to_float(row[absent_col])
                    total = (pres or 0) + (absv or 0)
                    val = round(100.0 * (absv or 0) / total, 2) if total else 0
                except (TypeError, ZeroDivisionError):
                    val = 0
            fill = (255, 255, 255)
            if ri == 0:
                fill = (54, 96, 146)
            elif ci in _COLOR_SCALE_COLS and ci < len(row):
                num = _value_to_float(val)
                if num is not None and col_mins[ci] is not None:
                    fill = _color_scale_rgb(num, col_mins[ci], col_maxs[ci] or col_mins[ci])
            if ri == 0:
                draw_cell(ri, ci, val, fill=fill, font_=font_header or font_bold, max_chars=22)
            else:
                draw_cell(ri, ci, val, fill=fill)

    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


def get_plant_report_send_time():
    """Return (hour, minute) 24h from SystemSetting plant_report_send_time (default 6:00)."""
    try:
        obj = SystemSetting.objects.get(key='plant_report_send_time')
        val = (obj.value or '06:00').strip()
        parts = val.split(':')
        h = int(parts[0]) if parts else 6
        m = int(parts[1]) if len(parts) > 1 else 0
        return (max(0, min(23, h)), max(0, min(59, m)))
    except (SystemSetting.DoesNotExist, ValueError, IndexError):
        return (6, 0)


def is_plant_report_email_enabled():
    try:
        obj = SystemSetting.objects.get(key='plant_report_enabled')
        return (obj.value or 'true').strip().lower() in ('true', '1', 'yes')
    except SystemSetting.DoesNotExist:
        return True


def get_plant_report_last_sent_date():
    """Return the date we last sent the report, or None."""
    try:
        obj = SystemSetting.objects.get(key='plant_report_last_sent')
        if obj.value:
            return date.fromisoformat(obj.value)
    except (SystemSetting.DoesNotExist, ValueError):
        pass
    return None


def set_plant_report_last_sent_date(sent_date):
    SystemSetting.objects.update_or_create(
        key='plant_report_last_sent',
        defaults={'value': sent_date.isoformat(), 'description': 'Last Plant Report email sent date'},
    )


def get_plant_report_maam_amount():
    """Return ma'am amount (final total salary as per ma'am) for difference calc, or None if not set."""
    try:
        obj = SystemSetting.objects.get(key='plant_report_maam_amount')
        v = (obj.value or '').strip()
        if not v:
            return None
        return float(v)
    except (SystemSetting.DoesNotExist, ValueError):
        return None


def build_plant_report_previous_day_excel_only():
    """
    Build an Excel file with only the Plant Report (Previous day) sheet.
    Same logic and columns as the Google Sheet tab. Starts at row 2, column 2 (B2).
    Color scale (green -> yellow -> red) on Average Salary/hr, Absenteeism %, OT bonus (hrs), OT bonus (rs).
    Returns BytesIO.
    """
    rows = _build_sheet3_data()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Plant Report (Previous day)'
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    # Start at row 2, column 2 (B2)
    start_row, start_col = 2, 2
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=val)
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font
    last_row = start_row + len(rows) - 1
    # Column indices (0-based from start_col): 6=H Avg Salary, 7=I Avg Salary/hr, 8=J Absenteeism %, 9=K Total Salary, 10=L OT bonus hrs, 11=M OT bonus rs
    # Color scale (green low -> yellow mid -> red high) for Average Salary/hr (I), Absenteeism % (J), OT bonus (hrs) (L), OT bonus (rs) (M)
    col_I = get_column_letter(start_col + 7)
    col_J = get_column_letter(start_col + 8)
    col_L = get_column_letter(start_col + 10)
    col_M = get_column_letter(start_col + 11)
    data_start_row = start_row + 1  # first data row below header
    range_ij = f'{col_I}{data_start_row}:{col_J}{last_row}'
    range_lm = f'{col_L}{data_start_row}:{col_M}{last_row}'
    # Muted colors: soft green -> pale yellow -> soft coral (less vibrant)
    color_scale = ColorScaleRule(
        start_type='min', start_color='C8E6C9',   # soft green
        mid_type='percentile', mid_value=50, mid_color='FFF9C4',  # pale yellow
        end_type='max', end_color='FFCCBC',      # soft coral
    )
    ws.conditional_formatting.add(range_ij, color_scale)
    ws.conditional_formatting.add(range_lm, color_scale)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def send_plant_report_email():
    """
    Build Plant Report (Previous day) Excel (only that sheet, same as Google Sheet) and email to all active recipients.
    Uses EmailSmtpConfig for SMTP. Returns dict with success, message, sent_count.
    """
    recipients = list(
        PlantReportRecipient.objects.filter(is_active=True).values_list('email', flat=True)
    )
    if not recipients:
        return {'success': False, 'message': 'No active recipients.', 'sent_count': 0}

    config = get_active_smtp_config()
    if not config or not config.auth_username:
        return {'success': False, 'message': 'SMTP not configured. Set in Settings / Django Admin.', 'sent_count': 0}

    try:
        buf = build_plant_report_previous_day_excel_only()
        excel_bytes = buf.read()
    except Exception as e:
        logger.exception('Failed to generate Plant Report Excel')
        return {'success': False, 'message': str(e), 'sent_count': 0}

    from_date = (timezone.localdate() - timezone.timedelta(days=1)).strftime('%d-%m-%Y')
    subject = f'Plant Report (Previous day) - {from_date}'
    sender_address = config.force_sender or config.auth_username
    from_email = f'Tubematic (Carrigar) <{sender_address}>' if sender_address else 'Tubematic (Carrigar) <noreply@tubematic.com>'

    # Our amount = Total Salary + OT bonus (rs) from Plant Report total row (previous day). Difference = ma'am amount - our amount.
    diff_block_text = ''
    diff_block_html = ''
    try:
        rows = _build_sheet3_data()
        if rows and len(rows) >= 2:
            total_row = rows[-1]
            n_dates = max(0, len(rows[0]) - 11)
            total_salary = _value_to_float(total_row[3 + n_dates + 5])
            ot_bonus_rs = _value_to_float(total_row[3 + n_dates + 7])
            our_amount = (total_salary or 0) + (ot_bonus_rs or 0)
            maam = get_plant_report_maam_amount()
            if maam is not None:
                diff = round(maam - our_amount, 2)
                diff_block_text = (
                    f'\n\nMa\'am amount: {maam:,.2f}  |  Our amount: {our_amount:,.2f}  |  Difference: {diff:,.2f}\n'
                )
                diff_block_html = (
                    '<p style="margin: 16px 0; font-size: 14px; padding: 12px; background: #f0f4f8; border-radius: 6px; border-left: 4px solid #366092;">'
                    f'<strong>Ma\'am amount:</strong> {maam:,.2f} &nbsp;|&nbsp; <strong>Our amount:</strong> {our_amount:,.2f} &nbsp;|&nbsp; <strong>Difference:</strong> {diff:,.2f}'
                    '</p>'
                )
    except Exception as e:
        logger.warning('Could not compute difference for Plant Report email: %s', e)

    # Build PNG image for mobile-friendly view
    img_bytes = None
    img_buf = build_plant_report_previous_day_image()
    if img_buf:
        img_bytes = img_buf.read()

    connection = get_connection(
        backend='django.core.mail.backends.smtp.EmailBackend',
        host=config.smtp_server,
        port=config.smtp_port,
        username=config.auth_username,
        password=config.auth_password,
        use_tls=True,
        fail_silently=False,
    )
    filename_xlsx = f'Plant_Report_{from_date.replace("-", "_")}.xlsx'
    text_body = (
        f'Dear Team,\n\nPlease find the Plant Report (Previous day) for {from_date}.\n\n'
        'The report is shown below for quick view on any device. The full Excel file is attached for download.'
        f'{diff_block_text}\n\n'
        'Regards,\nTubematic (Carrigar)'
    )
    if img_bytes:
        html_body = (
            '<div style="font-family: Arial, sans-serif; max-width: 640px; margin: 0 auto; color: #333;">'
            '<div style="padding: 16px 0 8px 0; border-bottom: 1px solid #e0e0e0;">'
            '<p style="margin: 0; font-size: 15px;">Dear Team,</p>'
            '</div>'
            f'<p style="margin: 16px 0; font-size: 14px; line-height: 1.5;">Please find the <strong>Plant Report (Previous day)</strong> for <strong>{from_date}</strong>.</p>'
            '<p style="margin: 12px 0; font-size: 14px; line-height: 1.5;">The report is shown below for quick view on any device. The full Excel file is attached for download.</p>'
            f'{diff_block_html}'
            '<div style="margin: 20px 0; padding: 12px 0; border: 1px solid #e8e8e8; border-radius: 6px; overflow-x: auto;">'
            '<img src="cid:plant_report" alt="Plant Report" style="max-width:100%; height:auto; display:block;" />'
            '</div>'
            '<div style="padding: 16px 0 0 0; border-top: 1px solid #e0e0e0; margin-top: 20px; font-size: 13px; color: #666;">'
            '<p style="margin: 0 0 4px 0;">Regards,</p>'
            '<p style="margin: 0; font-weight: 600;">Tubematic (Carrigar)</p>'
            '</div>'
            '</div>'
        )
        msg = EmailMultiAlternatives(subject, text_body, from_email, recipients, connection=connection)
        msg.attach_alternative(html_body, 'text/html')
        msg.mixed_subtype = 'related'
        msg_img = MIMEImage(img_bytes)
        msg_img.add_header('Content-ID', '<plant_report>')
        msg_img.add_header('Content-Disposition', 'inline', filename='Plant_Report.png')
        msg.attach(msg_img)
    else:
        msg = EmailMultiAlternatives(subject, text_body, from_email, recipients, connection=connection)
    msg.attach(filename_xlsx, excel_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    try:
        msg.send()
        set_plant_report_last_sent_date(timezone.localdate())
        logger.info('Plant Report email sent to %s recipients', len(recipients))
        return {'success': True, 'message': f'Sent to {len(recipients)} recipient(s).', 'sent_count': len(recipients)}
    except Exception as e:
        logger.exception('Plant Report email send failed')
        return {'success': False, 'message': str(e), 'sent_count': 0}


def maybe_send_plant_report_daily():
    """
    Call from a cron/task every minute (or every hour). If current time matches
    plant_report_send_time and we haven't sent today, send the report.
    """
    if not is_plant_report_email_enabled():
        return
    today = timezone.localdate()
    if get_plant_report_last_sent_date() == today:
        return
    hour, minute = get_plant_report_send_time()
    now = timezone.now()
    if now.hour == hour and now.minute == minute:
        send_plant_report_email()
