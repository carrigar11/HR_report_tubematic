"""
Generate payroll-style Excel export for Export Center.
Uses Django ORM. Date filter: month+year, single_date, or date_from/date_to.
previous_day: daily data = yesterday only, Total Salary = current month (1st through yesterday).
"""
from datetime import date, timedelta
from io import BytesIO

from django.db.models import Q, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Employee, Attendance, SalaryAdvance, Salary, Penalty


def _time_to_hours(t):
    if t is None:
        return None
    return t.hour + t.minute / 60.0 + t.second / 3600.0


def _shift_hours(shift_from, shift_to):
    if shift_from is None or shift_to is None:
        return 8.0
    f = _time_to_hours(shift_from)
    t = _time_to_hours(shift_to)
    if f is None or t is None:
        return 8.0
    if t <= f:
        t += 24.0
    return round(t - f, 2)


def _get_date_filter_queryset(date_from=None, date_to=None, single_date=None, month=None, year=None):
    """Return Attendance queryset filtered by the given date options."""
    qs = Attendance.objects.all()
    if single_date:
        qs = qs.filter(date=single_date)
    elif month is not None and year is not None:
        from calendar import monthrange
        _, last_day = monthrange(year, month)
        start = date(year, month, 1)
        end = date(year, month, last_day)
        qs = qs.filter(date__gte=start, date__lte=end)
    elif date_from or date_to:
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
    return qs


def _get_advance_by_emp(month=None, year=None, date_from=None, date_to=None, allowed_emp_codes=None):
    """Return dict emp_code -> total advance (float) for the given period."""
    qs = SalaryAdvance.objects.values('emp_code').annotate(total=Sum('amount'))
    if month is not None and year is not None:
        qs = qs.filter(month=month, year=year)
    elif date_from or date_to:
        from calendar import monthrange
        months_years = set()
        start = date_from or date(2000, 1, 1)
        end = date_to or date(2100, 12, 31)
        d = start.replace(day=1)
        while d <= end:
            months_years.add((d.month, d.year))
            _, last = monthrange(d.year, d.month)
            d = d.replace(day=last) + timedelta(days=1)
        if months_years:
            q = Q()
            for m, y in months_years:
                q = q | Q(month=m, year=y)
            qs = qs.filter(q)
    else:
        return {}
    if allowed_emp_codes is not None:
        qs = qs.filter(emp_code__in=allowed_emp_codes) if allowed_emp_codes else qs.none()
    return {r['emp_code']: float(r['total'] or 0) for r in qs}


def _get_penalty_by_emp(month=None, year=None, date_from=None, date_to=None, allowed_emp_codes=None):
    """Return dict emp_code -> total penalty deduction (float) for the given period."""
    qs = Penalty.objects.values('emp_code').annotate(total=Sum('deduction_amount'))
    if month is not None and year is not None:
        qs = qs.filter(month=month, year=year)
    elif date_from or date_to:
        from calendar import monthrange
        months_years = set()
        start = date_from or date(2000, 1, 1)
        end = date_to or date(2100, 12, 31)
        d = start.replace(day=1)
        while d <= end:
            months_years.add((d.month, d.year))
            _, last = monthrange(d.year, d.month)
            d = d.replace(day=last) + timedelta(days=1)
        if months_years:
            q = Q()
            for m, y in months_years:
                q = q | Q(month=m, year=y)
            qs = qs.filter(q)
    else:
        return {}
    if allowed_emp_codes is not None:
        qs = qs.filter(emp_code__in=allowed_emp_codes) if allowed_emp_codes else qs.none()
    return {r['emp_code']: float(r['total'] or 0) for r in qs}


def build_payroll_rows(employees, attendance_queryset, advance_by_emp=None, penalty_by_emp=None):
    """Build payroll matrix: one row per employee, date cols = daily earnings (rate × hours). advance_by_emp: dict emp_code -> advance amount. penalty_by_emp: dict emp_code -> penalty deduction."""
    if advance_by_emp is None:
        advance_by_emp = {}
    if penalty_by_emp is None:
        penalty_by_emp = {}
    # (emp_code, date) -> total_working_hours
    att_list = list(attendance_queryset.values('emp_code', 'date', 'total_working_hours'))
    att_map = {}
    dates_set = set()
    for r in att_list:
        key = (r['emp_code'], r['date'])
        h = r.get('total_working_hours')
        att_map[key] = float(h) if h is not None else 0.0
        dates_set.add(r['date'])
    sorted_dates = sorted(dates_set)

    payroll_rows = []
    for emp in employees:
        base = float(emp.base_salary or 0)
        salary_type = (emp.salary_type or 'Monthly').strip() or 'Monthly'
        if salary_type == 'Hourly':
            hourly_rate = base
        elif salary_type == 'Fixed':
            hourly_rate = 0.0
        else:
            hourly_rate = base / (26 * 8) if base else 0.0
        du = _shift_hours(emp.shift_from, emp.shift_to)
        advance_val = round(advance_by_emp.get(emp.emp_code, 0), 2)
        penalty_val = round(penalty_by_emp.get(emp.emp_code, 0), 2)

        row = {
            'emp_code': emp.emp_code,
            'name': emp.name or '',
            'pla': 'A',
            'status': emp.status or 'Working',
            'under_work': '',
            'department': emp.dept_name or '',
            'sala': round(hourly_rate, 2),
            'du': du,
            'advance': advance_val,
            'penalty': penalty_val,
        }
        day_totals = []
        row_total = 0.0
        for d in sorted_dates:
            hours = att_map.get((emp.emp_code, d), 0.0)
            amount = round(hourly_rate * hours, 2) if salary_type != 'Fixed' else 0.0
            day_totals.append(amount)
            row_total += amount
        if salary_type == 'Fixed':
            row_total = base
        else:
            row_total = round(row_total, 2)
        row['total'] = round(row_total, 2)
        row['_dates'] = sorted_dates
        row['_day_totals'] = day_totals
        payroll_rows.append(row)

    return sorted_dates, payroll_rows


def _add_bonus_to_payroll_rows(payroll_rows, month, year):
    """Add bonus (hours × hourly_rate) to each row's total. Also set row['bonus_hours'] and row['bonus_amount']."""
    if not payroll_rows or month is None or year is None:
        return
    emp_codes = [r['emp_code'] for r in payroll_rows]
    salary_type_by_emp = {e['emp_code']: (e.get('salary_type') or 'Monthly').strip() or 'Monthly' for e in Employee.objects.filter(emp_code__in=emp_codes).values('emp_code', 'salary_type')}
    bonus_by_emp = {}
    for s in Salary.objects.filter(emp_code__in=emp_codes, month=month, year=year).values('emp_code', 'bonus', 'base_salary'):
        bonus_by_emp[s['emp_code']] = (float(s.get('bonus') or 0), float(s.get('base_salary') or 0))
    for row in payroll_rows:
        ec = row['emp_code']
        sala = row.get('sala') or 0.0  # hourly rate already set in build_payroll_rows
        bonus_hrs, base = bonus_by_emp.get(ec, (0.0, 0.0))
        row['bonus_hours'] = round(bonus_hrs, 2)
        row['bonus_amount'] = 0.0
        if bonus_hrs <= 0:
            continue
        st = salary_type_by_emp.get(ec, 'Monthly')
        if st == 'Hourly':
            hourly_rate = base
        elif st == 'Fixed':
            hourly_rate = base / 208.0 if base else 0.0
        else:
            hourly_rate = base / 208.0 if base else 0.0
        bonus_money = round(bonus_hrs * hourly_rate, 2)
        row['bonus_amount'] = bonus_money
        row['total'] = round((row.get('total') or 0) + bonus_money, 2)


def _set_bonus_columns_for_date_range(payroll_rows, sorted_dates, add_bonus_to_total=True):
    """Set bonus_hours and bonus_amount on each row by summing bonus for all (month, year) in sorted_dates.
    If add_bonus_to_total True, also add bonus amount to row['total'] (so All dates / From–to Total Salary includes bonus)."""
    if not payroll_rows or not sorted_dates:
        return
    months_years = sorted(set((d.month, d.year) for d in sorted_dates if hasattr(d, 'month')))
    if not months_years:
        return
    emp_codes = [r['emp_code'] for r in payroll_rows]
    salary_type_by_emp = {
        e['emp_code']: (e.get('salary_type') or 'Monthly').strip() or 'Monthly'
        for e in Employee.objects.filter(emp_code__in=emp_codes).values('emp_code', 'salary_type')
    }
    # (emp_code, month, year) -> (bonus_hrs, base_salary)
    bonus_data = {}
    for (m, y) in months_years:
        for s in Salary.objects.filter(emp_code__in=emp_codes, month=m, year=y).values('emp_code', 'bonus', 'base_salary'):
            key = (s['emp_code'], m, y)
            bonus_data[key] = (float(s.get('bonus') or 0), float(s.get('base_salary') or 0))
    for row in payroll_rows:
        ec = row['emp_code']
        sala = row.get('sala') or 0.0
        st = salary_type_by_emp.get(ec, 'Monthly')
        total_hrs = 0.0
        total_amt = 0.0
        for (m, y) in months_years:
            bonus_hrs, base = bonus_data.get((ec, m, y), (0.0, 0.0))
            if bonus_hrs <= 0:
                continue
            if st == 'Hourly':
                hourly_rate = base
            elif st == 'Fixed':
                hourly_rate = base / 208.0 if base else 0.0
            else:
                hourly_rate = base / 208.0 if base else 0.0
            total_hrs += bonus_hrs
            total_amt += bonus_hrs * hourly_rate
        row['bonus_hours'] = round(total_hrs, 2)
        row['bonus_amount'] = round(total_amt, 2)
        if add_bonus_to_total and total_amt > 0:
            row['total'] = round((row.get('total') or 0) + total_amt, 2)


def write_payroll_sheet(ws, sorted_dates, payroll_rows, include_punch_columns=False, include_bonus_columns=False):
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['Emp Code', 'STAFF', 'Pla', 'Status', 'Under Work', 'Department', 'Sala', 'Du']
    if include_punch_columns:
        headers.extend(['Punch In', 'Punch Out'])
    for d in sorted_dates:
        headers.append(d.strftime('%d-%m-%y') if hasattr(d, 'strftime') else str(d))
    headers.append('TOTAL')
    headers.append('Advance')
    headers.append('Penalty')
    if include_bonus_columns:
        headers.extend(['Bonus (hrs)', 'Bonus (Rs)'])

    base_date_col = 11 if include_punch_columns else 9
    total_col = base_date_col + len(sorted_dates)
    advance_col = total_col + 1
    penalty_col = total_col + 2
    bonus_hrs_col = penalty_col + 1 if include_bonus_columns else None
    bonus_amt_col = penalty_col + 2 if include_bonus_columns else None

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font

    for row_idx, row in enumerate(payroll_rows, 2):
        ws.cell(row=row_idx, column=1, value=row.get('emp_code'))
        ws.cell(row=row_idx, column=2, value=row.get('name'))
        ws.cell(row=row_idx, column=3, value=row.get('pla'))
        ws.cell(row=row_idx, column=4, value=row.get('status'))
        ws.cell(row=row_idx, column=5, value=row.get('under_work'))
        ws.cell(row=row_idx, column=6, value=row.get('department'))
        ws.cell(row=row_idx, column=7, value=row.get('sala'))
        ws.cell(row=row_idx, column=8, value=row.get('du'))
        if include_punch_columns:
            ws.cell(row=row_idx, column=9, value=row.get('punch_in', ''))
            ws.cell(row=row_idx, column=10, value=row.get('punch_out', ''))
        for col_idx, day_val in enumerate(row.get('_day_totals', []), base_date_col):
            ws.cell(row=row_idx, column=col_idx, value=day_val)
        ws.cell(row=row_idx, column=total_col, value=row.get('total'))
        ws.cell(row=row_idx, column=advance_col, value=row.get('advance'))
        ws.cell(row=row_idx, column=penalty_col, value=row.get('penalty'))
        if include_bonus_columns:
            ws.cell(row=row_idx, column=bonus_hrs_col, value=row.get('bonus_hours'))
            ws.cell(row=row_idx, column=bonus_amt_col, value=row.get('bonus_amount'))

    tot_row = len(payroll_rows) + 2
    ws.cell(row=tot_row, column=1, value='Total').font = Font(bold=True)
    for col_idx in range(2, base_date_col):
        ws.cell(row=tot_row, column=col_idx, value='')
    col_totals = [0.0] * len(sorted_dates)
    grand_total = 0.0
    advance_total = 0.0
    penalty_total = 0.0
    bonus_total = 0.0
    for row in payroll_rows:
        for i, v in enumerate(row.get('_day_totals', [])):
            if i < len(col_totals):
                col_totals[i] += v
        grand_total += row.get('total') or 0
        advance_total += row.get('advance') or 0
        penalty_total += row.get('penalty') or 0
        bonus_total += row.get('bonus_amount') or 0
    for col_idx, tot in enumerate(col_totals, base_date_col):
        ws.cell(row=tot_row, column=col_idx, value=round(tot, 2)).font = Font(bold=True)
    ws.cell(row=tot_row, column=total_col, value=round(grand_total, 2)).font = Font(bold=True)
    ws.cell(row=tot_row, column=advance_col, value=round(advance_total, 2)).font = Font(bold=True)
    ws.cell(row=tot_row, column=penalty_col, value=round(penalty_total, 2)).font = Font(bold=True)
    if include_bonus_columns:
        ws.cell(row=tot_row, column=bonus_hrs_col, value='').font = Font(bold=True)
        ws.cell(row=tot_row, column=bonus_amt_col, value=round(bonus_total, 2)).font = Font(bold=True)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    if include_punch_columns:
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
    for col_idx in range(base_date_col, total_col):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    if include_bonus_columns:
        ws.column_dimensions[get_column_letter(bonus_hrs_col)].width = 12
        ws.column_dimensions[get_column_letter(bonus_amt_col)].width = 12


def build_plant_report_rows(payroll_rows, sorted_dates, attendance_queryset, month_total_per_dept=None):
    """
    One row per department (PLANT). Columns: Sr No, PLANT, Total Man Hrs, [date cols], Total Worker Present,
    Total Worker Absent, Average Salary, Average Salary/hr, Absenteeism %, Total Salary.
    If month_total_per_dept is provided (dept -> total), use it for Total Salary column; else use sum of date cols.
    """
    att_list = list(attendance_queryset.values('emp_code', 'date', 'total_working_hours', 'status'))
    emp_to_dept = {r['emp_code']: (r.get('department') or '') for r in payroll_rows}
    date_to_idx = {d: i for i, d in enumerate(sorted_dates)}

    # (dept, date) -> total_man_hrs, salary (from payroll), present_count, absent_count
    from collections import defaultdict
    dept_date_man_hrs = defaultdict(float)
    dept_date_salary = defaultdict(float)
    dept_date_present = defaultdict(int)
    dept_date_absent = defaultdict(int)

    for r in att_list:
        dept = emp_to_dept.get(r['emp_code'], '')
        dt = r['date']
        dept_date_man_hrs[(dept, dt)] += float(r.get('total_working_hours') or 0)
        if (r.get('status') or 'Present') == 'Present':
            dept_date_present[(dept, dt)] += 1
        else:
            dept_date_absent[(dept, dt)] += 1

    for row in payroll_rows:
        dept = row.get('department') or ''
        for i, d in enumerate(row.get('_dates', [])):
            amt = row.get('_day_totals', [])[i] if i < len(row.get('_day_totals', [])) else 0
            dept_date_salary[(dept, d)] += amt

    # Aggregate bonus by department (from payroll_rows: bonus from start-of-month till date or as per range)
    dept_bonus_hrs = defaultdict(float)
    dept_bonus_amount = defaultdict(float)
    # Total Salary per dept = sum of each employee's row total (includes Fixed salary + bonus; matches Payroll sheet)
    dept_total_from_rows = defaultdict(float)
    for row in payroll_rows:
        dept = row.get('department') or ''
        dept_bonus_hrs[dept] += float(row.get('bonus_hours') or 0)
        dept_bonus_amount[dept] += float(row.get('bonus_amount') or 0)
        dept_total_from_rows[dept] += float(row.get('total') or 0)

    depts = sorted(set(emp_to_dept.values()))
    dept_list = [d for d in depts if d]
    if '' in depts:
        dept_list.append('')

    plant_rows = []
    for sr, dept in enumerate(dept_list, 1):
        total_man_hrs = 0.0
        day_totals = []
        total_salary = 0.0
        total_present = 0
        total_absent = 0
        for d in sorted_dates:
            man_hrs = dept_date_man_hrs.get((dept, d), 0)
            sal = dept_date_salary.get((dept, d), 0)
            total_man_hrs += man_hrs
            day_totals.append(round(sal, 2))
            total_salary += sal
            total_present += dept_date_present.get((dept, d), 0)
            total_absent += dept_date_absent.get((dept, d), 0)

        avg_salary = round(total_salary / total_present, 2) if total_present else 0
        avg_salary_hr = round(total_salary / total_man_hrs, 2) if total_man_hrs else 0
        total_worker = total_present + total_absent
        absenteeism = round(100.0 * total_absent / total_worker, 2) if total_worker else 0
        # Total Salary: when month_total_per_dept provided (Single day) use it; else use sum of row totals per dept
        # (so All dates / Month&year / From–to include Fixed salary and bonus and match Payroll sheet TOTAL)
        if month_total_per_dept is not None:
            total_salary = month_total_per_dept.get(dept, 0)
        else:
            total_salary = dept_total_from_rows.get(dept, 0)
        day_salary_sum = sum(day_totals)  # for avg_salary (always day-based: salary to give that day / present)
        plant_rows.append({
            'sr': sr,
            'plant': dept if dept else 'No Dept',
            'total_man_hrs': round(total_man_hrs, 2),
            '_day_totals': day_totals,
            'total_salary': round(total_salary, 2),
            'total_present': total_present,
            'total_absent': total_absent,
            'avg_salary': round(day_salary_sum / total_present, 2) if total_present else 0,
            'avg_salary_hr': avg_salary_hr,
            'absenteeism': absenteeism,
            'total_bonus_hours': round(dept_bonus_hrs.get(dept, 0), 2),
            'total_bonus_amount': round(dept_bonus_amount.get(dept, 0), 2),
        })
    return plant_rows


def write_plant_report_sheet(ws, sorted_dates, plant_rows):
    """Write Plant Report: one row per department with Total Man Hrs, date cols, Present, Absent, Avg Salary, etc."""
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    headers = ['Sr No', 'PLANT', 'Total Man Hrs']
    for d in sorted_dates:
        headers.append(d.strftime('%d-%m-%y') if hasattr(d, 'strftime') else str(d))
    headers.extend(['Total Worker Present', 'Total Worker Absent', 'Average Salary', 'Average Salary/hr', 'Absenteeism %', 'Total Salary', 'Total Bonus (hrs)', 'Total Bonus (Rs)'])

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font

    for row_idx, row in enumerate(plant_rows, 2):
        ws.cell(row=row_idx, column=1, value=row['sr'])
        ws.cell(row=row_idx, column=2, value=row['plant'])
        ws.cell(row=row_idx, column=3, value=row['total_man_hrs'])
        for col_idx, v in enumerate(row['_day_totals'], 4):
            ws.cell(row=row_idx, column=col_idx, value=v)
        off = 4 + len(sorted_dates)
        ws.cell(row=row_idx, column=off, value=row['total_present'])
        ws.cell(row=row_idx, column=off + 1, value=row['total_absent'])
        ws.cell(row=row_idx, column=off + 2, value=row['avg_salary'])
        ws.cell(row=row_idx, column=off + 3, value=row['avg_salary_hr'])
        ws.cell(row=row_idx, column=off + 4, value=row['absenteeism'])
        ws.cell(row=row_idx, column=off + 5, value=row['total_salary'])
        ws.cell(row=row_idx, column=off + 6, value=row.get('total_bonus_hours', 0))
        ws.cell(row=row_idx, column=off + 7, value=row.get('total_bonus_amount', 0))

    # Total row
    tot_row = len(plant_rows) + 2
    ws.cell(row=tot_row, column=1, value='TOTAL SALARY').font = Font(bold=True)
    ws.cell(row=tot_row, column=2, value='')
    tot_man_hrs = sum(r['total_man_hrs'] for r in plant_rows)
    ws.cell(row=tot_row, column=3, value=round(tot_man_hrs, 2)).font = Font(bold=True)
    for col_idx in range(4, 4 + len(sorted_dates)):
        col_tot = sum(r['_day_totals'][col_idx - 4] for r in plant_rows)
        ws.cell(row=tot_row, column=col_idx, value=round(col_tot, 2)).font = Font(bold=True)
    off = 4 + len(sorted_dates)
    ws.cell(row=tot_row, column=off, value=sum(r['total_present'] for r in plant_rows)).font = Font(bold=True)
    ws.cell(row=tot_row, column=off + 1, value=sum(r['total_absent'] for r in plant_rows)).font = Font(bold=True)
    avg_overall = round(sum(r['total_salary'] for r in plant_rows) / sum(r['total_present'] for r in plant_rows), 2) if sum(r['total_present'] for r in plant_rows) else 0
    ws.cell(row=tot_row, column=off + 2, value=avg_overall).font = Font(bold=True)
    tot_sal = sum(r['total_salary'] for r in plant_rows)
    ws.cell(row=tot_row, column=off + 3, value=round(tot_sal / tot_man_hrs, 2) if tot_man_hrs else 0).font = Font(bold=True)
    ws.cell(row=tot_row, column=off + 4, value='').font = Font(bold=True)
    ws.cell(row=tot_row, column=off + 5, value=round(tot_sal, 2)).font = Font(bold=True)
    tot_bonus_hrs = sum(r.get('total_bonus_hours', 0) for r in plant_rows)
    tot_bonus_amt = sum(r.get('total_bonus_amount', 0) for r in plant_rows)
    ws.cell(row=tot_row, column=off + 6, value=round(tot_bonus_hrs, 2)).font = Font(bold=True)
    ws.cell(row=tot_row, column=off + 7, value=round(tot_bonus_amt, 2)).font = Font(bold=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    for col_idx in range(4, 4 + len(sorted_dates)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    for col_idx in range(off, off + 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def generate_payroll_excel_previous_day(allowed_emp_codes=None):
    """
    Report for previous day only: all daily data (date col, man hrs, present, absent, avg salary, avg/hr, absenteeism)
    for yesterday. Total Salary column = current month (1st through yesterday) for each employee/department.
    allowed_emp_codes: if set, only these employees (dept admin filter).
    """
    today = timezone.localdate()
    yesterday = today - timedelta(days=1)
    month_start = yesterday.replace(day=1)
    att_yesterday = _get_date_filter_queryset(single_date=yesterday)
    att_month = Attendance.objects.filter(date__gte=month_start, date__lte=yesterday)
    if allowed_emp_codes is not None:
        att_yesterday = att_yesterday.filter(emp_code__in=allowed_emp_codes)
        att_month = att_month.filter(emp_code__in=allowed_emp_codes)

    emp_qs = Employee.objects.all().order_by('dept_name', 'emp_code')
    if allowed_emp_codes is not None:
        emp_qs = emp_qs.filter(emp_code__in=allowed_emp_codes) if allowed_emp_codes else emp_qs.none()
    employees = list(emp_qs)
    advance_by_emp = _get_advance_by_emp(
        month=yesterday.month, year=yesterday.year,
        allowed_emp_codes=allowed_emp_codes
    )
    penalty_by_emp = _get_penalty_by_emp(
        month=yesterday.month, year=yesterday.year,
        allowed_emp_codes=allowed_emp_codes
    )
    sorted_dates, payroll_rows = build_payroll_rows(
        employees, att_yesterday, advance_by_emp=advance_by_emp, penalty_by_emp=penalty_by_emp
    )
    _, payroll_month = build_payroll_rows(
        employees, att_month, advance_by_emp=advance_by_emp, penalty_by_emp=penalty_by_emp
    )
    _add_bonus_to_payroll_rows(payroll_month, yesterday.month, yesterday.year)
    emp_to_month_total = {r['emp_code']: r['total'] for r in payroll_month}
    emp_to_bonus_hours = {r['emp_code']: r.get('bonus_hours', 0) for r in payroll_month}
    emp_to_bonus_amount = {r['emp_code']: r.get('bonus_amount', 0) for r in payroll_month}
    for row in payroll_rows:
        ec = row['emp_code']
        row['total'] = emp_to_month_total.get(ec, 0)
        row['bonus_hours'] = emp_to_bonus_hours.get(ec, 0)
        row['bonus_amount'] = emp_to_bonus_amount.get(ec, 0)

    # For previous day only: add punch in/out per employee (single day data)
    punch_map = {}
    for a in att_yesterday.values('emp_code', 'date', 'punch_in', 'punch_out'):
        key = (a['emp_code'], a['date'])
        pi = a.get('punch_in')
        po = a.get('punch_out')
        punch_map[key] = (
            pi.strftime('%H:%M') if pi else '',
            po.strftime('%H:%M') if po else ''
        )
    for row in payroll_rows:
        pin, pout = punch_map.get((row['emp_code'], yesterday), ('', ''))
        row['punch_in'] = pin
        row['punch_out'] = pout

    month_total_per_dept = {}
    for row in payroll_rows:
        dept = row.get('department') or ''
        month_total_per_dept[dept] = month_total_per_dept.get(dept, 0) + row['total']

    plant_rows = build_plant_report_rows(
        payroll_rows, sorted_dates, att_yesterday,
        month_total_per_dept=month_total_per_dept
    )
    return _write_payroll_workbook(
        sorted_dates, payroll_rows, plant_rows,
        include_punch_columns=True, include_bonus_columns=True
    )


def _write_payroll_workbook(sorted_dates, payroll_rows, plant_rows, include_punch_columns=False, include_bonus_columns=False):
    """Build workbook with Plant Report, Payroll, and per-dept sheets. include_punch_columns: add Punch In/Out (e.g. previous day). include_bonus_columns: add Bonus (hrs) and Bonus (Rs)."""
    wb = Workbook()
    ws_plant = wb.active
    ws_plant.title = 'Plant Report'
    write_plant_report_sheet(ws_plant, sorted_dates, plant_rows)
    ws = wb.create_sheet(title='Payroll')
    write_payroll_sheet(ws, sorted_dates, payroll_rows, include_punch_columns=include_punch_columns, include_bonus_columns=include_bonus_columns)
    depts = sorted(set(r.get('department') or '' for r in payroll_rows))
    dept_list = [d for d in depts if d]
    if '' in depts:
        dept_list.append('')
    for dept in dept_list:
        subset = [r for r in payroll_rows if (r.get('department') or '') == dept]
        if not subset:
            continue
        title = (dept if dept else 'No_Dept')[:31]
        title = ''.join(c for c in title if c not in r'\/:*?[]')
        ws_dept = wb.create_sheet(title=title)
        write_payroll_sheet(ws_dept, sorted_dates, subset, include_punch_columns=include_punch_columns, include_bonus_columns=include_bonus_columns)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_payroll_excel(date_from=None, date_to=None, single_date=None, month=None, year=None, allowed_emp_codes=None, emp_code=None):
    """
    Generate payroll Excel workbook. Returns BytesIO.
    Filter: single_date (one day), or month+year, or date_from/date_to (range), or all if none set.
    allowed_emp_codes: if set, only these employees (dept admin filter).
    emp_code: if set, only this single employee (overrides to one-emp export).
    """
    att_qs = _get_date_filter_queryset(
        date_from=date_from, date_to=date_to,
        single_date=single_date, month=month, year=year
    )
    if allowed_emp_codes is not None:
        att_qs = att_qs.filter(emp_code__in=allowed_emp_codes) if allowed_emp_codes else att_qs.none()
    if emp_code:
        att_qs = att_qs.filter(emp_code=emp_code)
    emp_qs = Employee.objects.all().order_by('dept_name', 'emp_code')
    if emp_code:
        emp_qs = emp_qs.filter(emp_code=emp_code)
    elif allowed_emp_codes is not None:
        emp_qs = emp_qs.filter(emp_code__in=allowed_emp_codes) if allowed_emp_codes else emp_qs.none()
    employees = list(emp_qs)
    if single_date:
        advance_by_emp = _get_advance_by_emp(
            month=single_date.month, year=single_date.year,
            allowed_emp_codes=allowed_emp_codes
        )
        penalty_by_emp = _get_penalty_by_emp(
            month=single_date.month, year=single_date.year,
            allowed_emp_codes=allowed_emp_codes
        )
    elif month is not None and year is not None:
        advance_by_emp = _get_advance_by_emp(
            month=month, year=year,
            allowed_emp_codes=allowed_emp_codes
        )
        penalty_by_emp = _get_penalty_by_emp(
            month=month, year=year,
            allowed_emp_codes=allowed_emp_codes
        )
    else:
        advance_by_emp = _get_advance_by_emp(
            date_from=date_from, date_to=date_to,
            allowed_emp_codes=allowed_emp_codes
        )
        penalty_by_emp = _get_penalty_by_emp(
            date_from=date_from, date_to=date_to,
            allowed_emp_codes=allowed_emp_codes
        )
    sorted_dates, payroll_rows = build_payroll_rows(
        employees, att_qs, advance_by_emp=advance_by_emp, penalty_by_emp=penalty_by_emp
    )
    if month is not None and year is not None:
        _add_bonus_to_payroll_rows(payroll_rows, month, year)
    elif single_date:
        _add_bonus_to_payroll_rows(payroll_rows, single_date.month, single_date.year)
    else:
        _set_bonus_columns_for_date_range(payroll_rows, sorted_dates)

    # Total Salary in Plant Report: All dates = sum of all cols; Month&year = whole month (sum of cols);
    # Single day = month-to-date (1st of month till selected date); From–to = sum of range cols.
    month_total_per_dept = None
    if single_date:
        month_start = single_date.replace(day=1)
        att_mtd = _get_date_filter_queryset(date_from=month_start, date_to=single_date)
        if allowed_emp_codes is not None:
            att_mtd = att_mtd.filter(emp_code__in=allowed_emp_codes) if allowed_emp_codes else att_mtd.none()
        if emp_code:
            att_mtd = att_mtd.filter(emp_code=emp_code)
        _, payroll_mtd = build_payroll_rows(
            employees, att_mtd, advance_by_emp=advance_by_emp, penalty_by_emp=penalty_by_emp
        )
        _add_bonus_to_payroll_rows(payroll_mtd, single_date.month, single_date.year)
        month_total_per_dept = {}
        for row in payroll_mtd:
            dept = row.get('department') or ''
            month_total_per_dept[dept] = month_total_per_dept.get(dept, 0) + (row.get('total') or 0)
        for dept in month_total_per_dept:
            month_total_per_dept[dept] = round(month_total_per_dept[dept], 2)

    plant_rows = build_plant_report_rows(
        payroll_rows, sorted_dates, att_qs, month_total_per_dept=month_total_per_dept
    )
    return _write_payroll_workbook(
        sorted_dates, payroll_rows, plant_rows,
        include_bonus_columns=True
    )
