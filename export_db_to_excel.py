#!/usr/bin/env python3
"""
Standalone script (no Django). Exports all HR database tables to Excel:
one sheet per table with columns and data as stored in the database.

Run from project root:
  pip install psycopg2-binary openpyxl
  python export_db_to_excel.py

Uses same DB as Django: set DB_NAME, DB_USER, DB_PASSWORD, DB_HOST, DB_PORT
in backend/.env or environment (no default password).
"""

import json
import os
import re
from datetime import date, datetime
from pathlib import Path

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    print("Install: pip install psycopg2-binary")
    raise

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install: pip install openpyxl")
    raise


# Database connection (same as Django settings). Load from backend/.env if present.
def _load_env():
    env_path = Path(__file__).resolve().parent / "backend" / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                k, v = k.strip(), v.strip().strip('"').strip("'")
                if k and k not in os.environ:
                    os.environ[k] = v


_load_env()
DB_NAME = os.environ.get("DB_NAME", "hr_attendance_db")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "Seeta#4597")
DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_PORT = os.environ.get("DB_PORT", "5432")

# Tables to export: (table_name, ORDER BY clause). Sheet name = table name (sanitized).
# If FETCH_ALL_TABLES is True, all tables in public schema are discovered and exported.
FETCH_ALL_TABLES = True

TABLE_EXPORT_ORDER = [
    ("companies", "id"),
    ("company_registration_requests", "created_at DESC"),
    ("admins", "id"),
    ("employees", "dept_name, emp_code"),
    ("attendance", "date DESC, emp_code"),
    ("salaries", "year DESC, month DESC, emp_code"),
    ("salary_advances", "year DESC, month DESC, emp_code"),
    ("adjustment", "created_at DESC"),
    ("shift_overtime_bonus", "date DESC, emp_code"),
    ("penalty", "date DESC, emp_code"),
    ("penalty_inquiries", "created_at DESC"),
    ("performance_rewards", "created_at DESC"),
    ("holidays", "date"),
    ("leave_requests", "requested_at DESC"),
    ("system_settings", "key"),
    ("plant_report_recipients", "id"),
    ("email_smtp_config", "id"),
    ("audit_log", "created_at DESC"),
]


def sanitize_sheet_name(name):
    """Excel sheet names: max 31 chars, no \\ / : * ? [ ]"""
    if not name or not str(name).strip():
        return "No_Dept"
    s = re.sub(r'[\\/:*?\[\]]', '_', str(name).strip())
    return (s[:31]) if len(s) > 31 else s


def get_connection():
    if not DB_PASSWORD and not os.environ.get("DB_PASSWORD"):
        raise SystemExit(
            "DB_PASSWORD not set. Create backend/.env from backend/.env.example with DB_* values, or set DB_PASSWORD in the environment."
        )
    return psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT,
        cursor_factory=RealDictCursor,
    )


def get_all_tables(conn):
    """Return list of (table_name, order_by) for all tables in public schema. order_by = 'id' or first column."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT table_name FROM information_schema.tables
            WHERE table_schema = 'public' AND table_type = 'BASE TABLE'
            ORDER BY table_name
        """)
        tables = [r["table_name"] for r in cur.fetchall()]
    result = []
    for table_name in tables:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_schema = 'public' AND table_name = %s
                ORDER BY ordinal_position
            """, (table_name,))
            cols = [r["column_name"] for r in cur.fetchall()]
        if not cols:
            result.append((table_name, "1"))
            continue
        # Prefer id, then first column (quoted)
        if "id" in cols:
            order_by = "id"
        else:
            order_by = f'"{cols[0]}"'
        result.append((table_name, order_by))
    return result


def fetch_table(conn, table_name, order_by):
    """Fetch all rows from a table. Returns list of dicts; columns = keys from first row."""
    with conn.cursor() as cur:
        cur.execute(f'SELECT * FROM "{table_name}" ORDER BY {order_by}')
        return cur.fetchall()


def fetch_attendance_for_payroll(conn):
    """Emp code, date, total_working_hours for daily earnings calc (computed payroll)."""
    sql = """
    SELECT emp_code, date, total_working_hours
    FROM attendance
    ORDER BY date, emp_code
    """
    with conn.cursor() as cur:
        cur.execute(sql)
        return cur.fetchall()


def value_for_excel(v):
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.isoformat() if hasattr(v, "isoformat") else str(v)
    if hasattr(v, "isoformat"):  # time
        return str(v)[:5] if v else ""
    if isinstance(v, (dict, list)):
        return json.dumps(v) if v else ""
    return v


def _time_to_hours(t):
    """Convert time to decimal hours (e.g. 09:30 -> 9.5)."""
    if t is None:
        return None
    return t.hour + t.minute / 60.0 + t.second / 3600.0


def _shift_hours(shift_from, shift_to):
    """Expected duty hours per day from shift. If span next day, 24 - from + to."""
    if shift_from is None or shift_to is None:
        return 8.0
    f = _time_to_hours(shift_from)
    t = _time_to_hours(shift_to)
    if f is None or t is None:
        return 8.0
    if t <= f:
        t += 24.0
    return round(t - f, 2)


def build_payroll_data(emp_rows, att_rows):
    """
    Build payroll matrix: one row per employee, date columns = daily earnings
    (hourly_rate * total_working_hours that day). TOTAL = sum of date cols. Advance = 0.
    Returns (list of sorted dates, list of payroll rows, emp_code -> row index).
    """
    from decimal import Decimal
    # (emp_code, date) -> total_working_hours
    att_map = {}
    dates_set = set()
    for r in att_rows:
        key = (r["emp_code"], r["date"])
        h = r.get("total_working_hours")
        att_map[key] = float(h) if h is not None else 0.0
        dates_set.add(r["date"])
    sorted_dates = sorted(dates_set)

    emp_by_code = {r["emp_code"]: r for r in emp_rows}

    payroll_rows = []
    for emp in emp_rows:
        ec = emp["emp_code"]
        base = emp.get("base_salary")
        if base is not None:
            base = float(base)
        else:
            base = 0.0
        salary_type = (emp.get("salary_type") or "").strip() or "Monthly"
        # Hourly rate: Monthly -> base/(26*8), Hourly -> base
        if salary_type == "Hourly":
            hourly_rate = base
        else:
            hourly_rate = base / (26 * 8) if base else 0.0

        du = _shift_hours(emp.get("shift_from"), emp.get("shift_to"))

        row = {
            "emp_code": ec,
            "name": emp.get("name") or "",
            "pla": "A",
            "status": emp.get("status") or "Working",
            "under_work": "",
            "department": emp.get("dept_name") or "",
            "sala": round(hourly_rate, 2),
            "du": du,
            "advance": 0,
        }
        day_totals = []
        row_total = 0.0
        for d in sorted_dates:
            hours = att_map.get((ec, d), 0.0)
            amount = round(hourly_rate * hours, 2)
            row["date_" + d.isoformat()] = amount
            day_totals.append(amount)
            row_total += amount
        row["total"] = round(row_total, 2)
        row["_dates"] = sorted_dates
        row["_day_totals"] = day_totals
        payroll_rows.append(row)

    return sorted_dates, payroll_rows


def write_sheet_from_rows(ws, title, rows, columns):
    """Write header and rows to worksheet. columns = list of (key, header_label)."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (_, header) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, row in enumerate(rows, 2):
        for col_idx, (key, _) in enumerate(columns, 1):
            val = row.get(key)
            ws.cell(row=row_idx, column=col_idx, value=value_for_excel(val))
    # Auto-fit idea: set column width from max length
    for col_idx, (key, header) in enumerate(columns, 1):
        max_len = len(str(header))
        for r in rows:
            max_len = max(max_len, len(str(value_for_excel(r.get(key)))))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 1, 50)


def _column_header(col_name):
    """Human-friendly header from column name (e.g. emp_code -> Emp Code)."""
    return str(col_name).replace("_", " ").strip().title()


def write_sheet_from_table(ws, table_name, rows):
    """Write one sheet from table rows (list of dicts). Columns = keys from first row."""
    if not rows:
        ws.cell(row=1, column=1, value=f"(no data in {table_name})")
        return
    columns = [(k, _column_header(k)) for k in rows[0].keys()]
    write_sheet_from_rows(ws, table_name, rows, columns)


def write_payroll_sheet(ws, sorted_dates, payroll_rows, sheet_title):
    """
    Write payroll layout: Emp Code, STAFF, Pla, Status, Under Work, Department, Sala, Du,
    then one column per date (daily earnings), TOTAL, Advance. Last row = totals.
    """
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    # Headers
    headers = [
        "Emp Code", "STAFF", "Pla", "Status", "Under Work", "Department", "Sala", "Du"
    ]
    for i, d in enumerate(sorted_dates):
        headers.append(d.strftime("%d-%m-%y") if hasattr(d, "strftime") else str(d))
    headers.append("TOTAL")
    headers.append("Advance")

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font

    # Data rows
    for row_idx, row in enumerate(payroll_rows, 2):
        ws.cell(row=row_idx, column=1, value=row.get("emp_code"))
        ws.cell(row=row_idx, column=2, value=row.get("name"))
        ws.cell(row=row_idx, column=3, value=row.get("pla"))
        ws.cell(row=row_idx, column=4, value=row.get("status"))
        ws.cell(row=row_idx, column=5, value=row.get("under_work"))
        ws.cell(row=row_idx, column=6, value=row.get("department"))
        ws.cell(row=row_idx, column=7, value=row.get("sala"))
        ws.cell(row=row_idx, column=8, value=row.get("du"))
        for col_idx, day_val in enumerate(row.get("_day_totals", []), 9):
            ws.cell(row=row_idx, column=col_idx, value=day_val)
        ws.cell(row=row_idx, column=9 + len(sorted_dates), value=row.get("total"))
        ws.cell(row=row_idx, column=10 + len(sorted_dates), value=row.get("advance"))

    # Totals row at bottom
    tot_row = len(payroll_rows) + 2
    ws.cell(row=tot_row, column=1, value="Total")
    ws.cell(row=tot_row, column=1).font = Font(bold=True)
    for col_idx in range(2, 9):
        ws.cell(row=tot_row, column=col_idx, value="")
    col_totals = [0.0] * len(sorted_dates)
    grand_total = 0.0
    advance_total = 0.0
    for row in payroll_rows:
        for i, v in enumerate(row.get("_day_totals", [])):
            if i < len(col_totals):
                col_totals[i] += v
        grand_total += row.get("total") or 0
        advance_total += row.get("advance") or 0
    for col_idx, tot in enumerate(col_totals, 9):
        ws.cell(row=tot_row, column=col_idx, value=round(tot, 2))
        ws.cell(row=tot_row, column=col_idx).font = Font(bold=True)
    ws.cell(row=tot_row, column=9 + len(sorted_dates), value=round(grand_total, 2))
    ws.cell(row=tot_row, column=9 + len(sorted_dates)).font = Font(bold=True)
    ws.cell(row=tot_row, column=10 + len(sorted_dates), value=round(advance_total, 2))
    ws.cell(row=tot_row, column=10 + len(sorted_dates)).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    for col_idx in range(9, 9 + len(sorted_dates)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


def main():
    out_path = Path(__file__).resolve().parent / "hr_export_all_data.xlsx"
    print(f"Connecting to DB {DB_NAME}@{DB_HOST}...")
    conn = get_connection()
    stored = {}

    try:
        wb = Workbook()
        first_sheet = True

        tables_to_export = get_all_tables(conn) if FETCH_ALL_TABLES else TABLE_EXPORT_ORDER

        # One sheet per table: columns and data as in the database
        for table_name, order_by in tables_to_export:
            print(f"Fetching table: {table_name}...")
            rows = fetch_table(conn, table_name, order_by)
            stored[table_name] = rows
            print(f"  Got {len(rows)} rows.")
            sheet_title = sanitize_sheet_name(table_name)
            if first_sheet:
                ws = wb.active
                ws.title = sheet_title
                first_sheet = False
            else:
                ws = wb.create_sheet(title=sheet_title)
            write_sheet_from_table(ws, table_name, rows)
            print(f"  Written sheet: {sheet_title}")

        # Optional: computed Payroll sheets (daily earnings from employees + attendance)
        emp_rows = stored.get("employees", [])
        if emp_rows:
            print("Fetching attendance for payroll (computed view)...")
            att_payroll = fetch_attendance_for_payroll(conn)
            sorted_dates, payroll_rows = build_payroll_data(emp_rows, att_payroll)
            print(f"  Dates: {len(sorted_dates)} days, Employees: {len(payroll_rows)}")
            ws_payroll_all = wb.create_sheet(title="Payroll_All")
            write_payroll_sheet(ws_payroll_all, sorted_dates, payroll_rows, "Payroll_All")
            print("  Written sheet: Payroll_All (computed daily earnings)")
            depts = sorted(set(r.get("department") or "" for r in payroll_rows))
            dept_list = [d for d in depts if d]
            if any(not d for d in depts):
                dept_list.append("")
            for dept in dept_list:
                subset = [r for r in payroll_rows if (r.get("department") or "") == dept]
                if not subset:
                    continue
                sheet_name = "Payroll_" + sanitize_sheet_name(dept if dept else "No_Dept")
                ws = wb.create_sheet(title=sheet_name)
                write_payroll_sheet(ws, sorted_dates, subset, sheet_name)
                print(f"  Written sheet: {sheet_name} ({len(subset)} rows)")

        wb.save(out_path)
        print(f"\nSaved: {out_path}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
